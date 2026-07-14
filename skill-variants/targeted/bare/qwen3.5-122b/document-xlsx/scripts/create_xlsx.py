#!/usr/bin/env python3
"""
Bundled script for creating Excel spreadsheets with openpyxl.
Usage: python3 scripts/create_xlsx.py

This script is a TEMPLATE. Before running, edit the DATA and CONFIG sections
to match your specific task. The structure handles:
- Data entry with headers
- SUM and AVERAGE formulas
- Conditional formatting (red fill for cells below a threshold)
- Bold headers and currency formatting

IMPORTANT: Save to the EXACT filename requested by the user.
"""

import sys
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment

# ============================================================
# CONFIG - Edit these values for your task
# ============================================================
OUTPUT_FILE = "output.xlsx"  # CHANGE to the exact filename requested
SHEET_TITLE = "Sheet1"
THRESHOLD = 50000  # Cells below this get red fill

# ============================================================
# DATA - Replace with your actual data
# ============================================================
HEADERS = ["Product", "Q1", "Q2", "Q3", "Q4"]
DATA = [
    ["Widgets", 62000, 71000, 58000, 80000],
    ["Gadgets", 45000, 52000, 48000, 63000],
]

# ============================================================
# BUILD SPREADSHEET - Modify structure as needed
# ============================================================
wb = Workbook()
ws = wb.active
ws.title = SHEET_TITLE

# Write headers (row 1)
for col_idx, header in enumerate(HEADERS, 1):
    cell = ws.cell(row=1, column=col_idx, value=header)
    cell.font = Font(bold=True)

# Write data rows (starting row 2)
data_start_row = 2
for row_idx, row_data in enumerate(DATA, data_start_row):
    for col_idx, value in enumerate(row_data, 1):
        ws.cell(row=row_idx, column=col_idx, value=value)

data_end_row = data_start_row + len(DATA) - 1
num_data_cols = len(HEADERS) - 1  # exclude label column

# Add SUM row
sum_row = data_end_row + 1
ws.cell(row=sum_row, column=1, value="Total")
ws.cell(row=sum_row, column=1).font = Font(bold=True)
for col_idx in range(2, 2 + num_data_cols):
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    ws.cell(row=sum_row, column=col_idx,
            value=f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})")

# Add AVERAGE row
avg_row = sum_row + 1
ws.cell(row=avg_row, column=1, value="Average")
ws.cell(row=avg_row, column=1).font = Font(bold=True)
for col_idx in range(2, 2 + num_data_cols):
    col_letter = ws.cell(row=1, column=col_idx).column_letter
    ws.cell(row=avg_row, column=col_idx,
            value=f"=AVERAGE({col_letter}{data_start_row}:{col_letter}{data_end_row})")

# Apply red fill to data cells below threshold
red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
for row_idx in range(data_start_row, data_end_row + 1):
    for col_idx in range(2, 2 + num_data_cols):
        cell = ws.cell(row=row_idx, column=col_idx)
        if isinstance(cell.value, (int, float)) and cell.value < THRESHOLD:
            cell.fill = red_fill

# Apply currency format to all numeric cells
for row_idx in range(data_start_row, avg_row + 1):
    for col_idx in range(2, 2 + num_data_cols):
        ws.cell(row=row_idx, column=col_idx).number_format = "$#,##0"

# Auto-size columns
for col_idx in range(1, len(HEADERS) + 1):
    ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 15

# Save
wb.save(OUTPUT_FILE)
print(f"Created {OUTPUT_FILE} successfully")
