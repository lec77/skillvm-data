#!/usr/bin/env python3
"""Verify sales_report.xlsx structure, formulas, and formatting."""
import json
import sys
import os

def verify(xlsx_path):
    results = {
        "exists": False,
        "valid": False,
        "product_lines": [],
        "quarter_headers": [],
        "revenue_values": [],
        "sum_formulas": [],
        "average_formulas": [],
        "has_red_formatting": False,
        "has_conditional_formatting": False,
        "all_formulas": [],
        "error": None,
    }

    if not os.path.exists(xlsx_path):
        results["error"] = "File not found"
        return results

    results["exists"] = True

    try:
        from openpyxl import load_workbook

        # Check formulas (data_only=False)
        wb = load_workbook(xlsx_path, data_only=False)
        ws = wb.active
        results["valid"] = True

        # Scan all cells for labels, values, and formulas
        all_text = []
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
            for cell in row:
                val = cell.value
                if val is None:
                    continue
                sval = str(val)
                all_text.append(sval.lower())
                if isinstance(val, str) and val.startswith("="):
                    results["all_formulas"].append(val)
                    upper = val.upper()
                    if "SUM" in upper:
                        results["sum_formulas"].append(val)
                    if "AVERAGE" in upper:
                        results["average_formulas"].append(val)

        # Check product lines
        expected_products = ["widgets", "gadgets", "services", "licenses", "support"]
        for p in expected_products:
            if any(p in t for t in all_text):
                results["product_lines"].append(p)

        # Check quarter headers
        for q in ["q1", "q2", "q3", "q4"]:
            if any(q in t for t in all_text):
                results["quarter_headers"].append(q)

        # Check revenue values (read calculated values)
        wb_data = load_workbook(xlsx_path, data_only=True)
        ws_data = wb_data.active
        expected_revenues = [
            62000, 71000, 58000, 80000,  # Widgets
            45000, 52000, 48000, 63000,  # Gadgets
            38000, 42000, 35000, 47000,  # Services
            85000, 91000, 78000, 95000,  # Licenses
            28000, 32000, 25000, 36000,  # Support
        ]
        found_values = set()
        for row in ws_data.iter_rows():
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    found_values.add(int(cell.value))
        for rv in expected_revenues:
            if rv in found_values:
                results["revenue_values"].append(rv)
        wb_data.close()

        # Check for red formatting (conditional or direct)
        # Check conditional formatting rules
        if ws.conditional_formatting._cf_rules:
            results["has_conditional_formatting"] = True

        # Check direct cell fills/fonts for red-ish colors
        # Accept any red variant: pure red, dark red, light red, salmon, etc.
        def is_reddish(rgb_str):
            s = rgb_str.upper()
            # Common red hex patterns in openpyxl (AARRGGBB or RRGGBB)
            red_patterns = [
                "FF0000", "FF4444", "FF3333", "FF6666",  # Pure/bright reds
                "CC0000", "DD0000", "EE0000", "AA0000",  # Dark reds
                "FFC7CE", "F4CCCC", "FFD9D9", "FF9999",  # Light/pastel reds
                "C00000", "B00000", "990000",             # Very dark reds
            ]
            for pat in red_patterns:
                if pat in s:
                    return True
            # Heuristic: if R channel dominates G and B channels
            # Format is typically AARRGGBB (8 chars) or RRGGBB (6 chars)
            if len(s) >= 6:
                try:
                    h = s[-6:]
                    r, g, b = int(h[0:2], 16), int(h[2:4], 16), int(h[4:6], 16)
                    # Red-dominant: R is high and significantly above G and B
                    if r > 180 and r > g + 40 and r > b + 40:
                        return True
                except (ValueError, IndexError):
                    pass
            return False

        for row in ws.iter_rows():
            for cell in row:
                # Check font color
                if cell.font and cell.font.color and cell.font.color.rgb:
                    rgb = str(cell.font.color.rgb)
                    if is_reddish(rgb):
                        results["has_red_formatting"] = True
                # Check fill color
                if cell.fill and cell.fill.fgColor and cell.fill.fgColor.rgb:
                    rgb = str(cell.fill.fgColor.rgb)
                    if is_reddish(rgb):
                        results["has_red_formatting"] = True

        wb.close()

    except Exception as e:
        results["error"] = str(e)

    return results


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "sales_report.xlsx"
    print(json.dumps(verify(path), indent=2))
