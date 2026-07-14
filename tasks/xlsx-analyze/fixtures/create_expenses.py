#!/usr/bin/env python3
"""Generate expenses.xlsx fixture for the xlsx-analyze benchmark task."""

import sys
import os

try:
    import openpyxl
except ImportError:
    print("Installing openpyxl...", file=sys.stderr)
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    import openpyxl

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Expenses"

# Header row
ws.append(["Month", "Category", "Amount", "Approved"])

# Deterministic expense data: 4 categories x 6 months = 24 rows
# Travel:    Jan=1200, Feb=800,  Mar=1500, Apr=950,  May=1100, Jun=1350  -> 6900
# Software:  Jan=500,  Feb=500,  Mar=2000, Apr=500,  May=500,  Jun=3000  -> 7000
# Equipment: Jan=3500, Feb=0,    Mar=800,  Apr=0,    May=4200, Jun=0     -> 8500
# Training:  Jan=1500, Feb=2000, Mar=1500, Apr=2500, May=1500, Jun=2000  -> 11000
# Grand total: 33400
# Approved: 18 of 24 -> 75%

data = [
    # (Month, Category, Amount, Approved)
    ("Jan", "Travel",    1200, True),
    ("Jan", "Software",   500, True),
    ("Jan", "Equipment", 3500, True),
    ("Jan", "Training",  1500, True),
    ("Feb", "Travel",     800, True),
    ("Feb", "Software",   500, True),
    ("Feb", "Equipment",    0, False),
    ("Feb", "Training",  2000, True),
    ("Mar", "Travel",    1500, True),
    ("Mar", "Software",  2000, True),
    ("Mar", "Equipment",  800, False),
    ("Mar", "Training",  1500, True),
    ("Apr", "Travel",     950, False),
    ("Apr", "Software",   500, True),
    ("Apr", "Equipment",    0, False),
    ("Apr", "Training",  2500, True),
    ("May", "Travel",    1100, True),
    ("May", "Software",   500, False),
    ("May", "Equipment", 4200, True),
    ("May", "Training",  1500, True),
    ("Jun", "Travel",    1350, True),
    ("Jun", "Software",  3000, True),
    ("Jun", "Equipment",    0, False),
    ("Jun", "Training",  2000, True),
]

for row in data:
    ws.append(list(row))

# Add SUM formula row at the bottom (row 26 = after header + 24 data rows)
sum_row = 26
ws.cell(row=sum_row, column=1, value="TOTAL")
ws.cell(row=sum_row, column=3).value = "=SUM(C2:C25)"

output_path = sys.argv[1] if len(sys.argv) > 1 else "expenses.xlsx"
wb.save(output_path)
print(f"Created {output_path} with {len(data)} expense rows")
